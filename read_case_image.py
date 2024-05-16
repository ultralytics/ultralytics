import argparse
import json
import os
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from PIL import Image

parser = argparse.ArgumentParser(description="读取案件图片")
parser.add_argument("path", help="path of excel file")
args = parser.parse_args()


def parser_merged_cell(sheet, row, col):
    """检查是否为合并单元格并获取对应行列单元格的值。 如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值 :param sheet: 当前工作表对象 :param row:
    需要获取的单元格所在行 :param col: 需要获取的单元格所在列 :return:
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell


if __name__ == "__main__":
    cellnames = []

    root = args.path
    urls = []
    for workbook in os.listdir(root):
        if not workbook.endswith(".xlsx"):
            continue
        wb = load_workbook(os.path.join(root, workbook))
        ws = wb[wb.sheetnames[0]]
        for row_index in range(1, 2):
            for col_index in range(1, ws.max_column + 1):
                cell_ = parser_merged_cell(ws, row_index, col_index)
                # cell_ = ws.cell(row=row_index, column=col_index)
                cellnames.append(cell_.value)

        images = ws._images
        for row_index in range(3, ws.max_row + 1):
            code = ""
            imgs = []
            for col_index in range(1, ws.max_column + 1):
                if cellnames[col_index - 1] == "检查图片(上报不得少于三张照片:近景，远景，标志物)":
                    # code = ws.cell(row=row_index, column=col_index).value
                    # imgs.append(ws.cell(row=row_index, column=col_index).value)
                    # break
                    value = ws.cell(row=row_index, column=col_index).value
                    if value is not None:
                        if value.startswith("https://watermark.survey.work"):
                            url = urlparse(url=value)
                            value = unquote(url.path.split("/")[-1])
                        value = value.replace("ow-prod-cdn.survey.work", "ow-prod.oss-cn-beijing.aliyuncs.com")
                        path = urlparse(url=value).path
                        print(path)
                        urls.append(path)
    json.dump(urls, open(os.path.join(root, "urls.json"), "w"))

    # imgs = [Image.open(img.ref).convert('RGB') for img in images if img.anchor._from.row == row_index - 1 and cellnames[img.anchor._from.col] == '检查图片(上报不得少于三张照片:近景，远景，标志物)	检查图片(上报不得少于三张照片:近景，远景，标志物)']
    # code_path = os.path.join(args.result_path, code)
    # if not os.path.exists(code_path):
    #     os.makedirs(code_path)
    # for idx, img in enumerate(imgs):
    #     img.save(os.path.join(code_path, f'IMG_{idx}.jpg'))
